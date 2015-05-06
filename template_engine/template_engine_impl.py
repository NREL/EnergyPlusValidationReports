#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import re
import yaml
import datetime
import os
import codecs
import numpy as np
import math

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import openpyxl
import sys
import urllib
import textwrap
import time
import csv
import collections
import shutil
import logging
import functools

from matplotlib import cm

from openpyxl import load_workbook, worksheet, cell

from functools import partial



logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')

engine = None

class FileLookupException(Exception):
    def __init__(self, filename):
        super(FileLookupException, self).__init__("File lookup failure: " + filename)
        self.filename = filename

def eval_replacement(matchobj):
    try:
        return str(eval(matchobj.group(1)))
    except:
        logging.error("Error while evaling: '" + matchobj.group(1) + "'")
        raise

def flatten(l):
    for el in l:
        if isinstance(el, collections.Iterable) and not isinstance(el, str):
            for sub in flatten(el):
                yield sub
        else:
            yield el

def non_empty_floats(values):
    retval = []

    for idx, v in enumerate(flatten(values)):
        try:
            if isinstance(v, bool):
                pass
            elif isinstance(v, float):
                retval.append(v)
            elif isinstance(v, int):
                retval.append(float(v))
            else:
                if v != None and v.strip() != "":
                    retval.append(float(v))
        except ValueError:
            logging.warning("Error with float conversion in non_empty_floats, index " + str(idx) + " val: " + str(v) )

    return retval

def gettime(s):
    #logging.debug("gettime: " + str(s))
    val = time.strptime(s.strip(), "%m/%d  %H:%M:%S")
    #logging.debug("val: " + str(val))
    t = datetime.time(val.tm_hour, val.tm_min)
    #logging.debug("wastime: " + str(t))
    return t

def istime(s):

    try:
        gettime(s)
        return True
    except:
        return False

def normalize_headings(sheet_headings, user_headings):

    is_times = any(any(isinstance(h, datetime.time) for h in heading) for heading in sheet_headings)

    is_times = is_times or all(all(istime(h) for h in heading) for heading in sheet_headings)


    normalized_headings = sheet_headings

#    logging.debug("is_times: " + str(is_times))

    if is_times:
        normalized_headings = []

        for heading in sheet_headings:
            normalized_heading = []
            for h in heading:
#                logging.debug("is_times: " + str(h))
                if not isinstance(h, datetime.time):
                    try:
                        dt = (datetime.datetime(1900,1,1) + datetime.timedelta(0, 60*60*24*float(h))).time()
                        normalized_heading.append(datetime.time(dt.hour, dt.minute))
                    except:
#                        logging.debug("except: " + str(h))
                        try:
                            normalized_heading.append(gettime(h))
                        except:
                            normalized_heading.append(datetime.time(0,0))
                else:
                    normalized_heading.append(h)
            normalized_headings.append(normalized_heading)


    xs = normalized_headings

    if is_times:
        xs = [[time.hour * 60 + time.minute for time in heading] for heading in normalized_headings]


    # print(normalized_headings)

    filtered_headings = set(normalized_headings[0]).intersection(*normalized_headings)
    if is_times:
        filtered_headings.discard(datetime.time(0,0))

    #print(str(filtered_headings))

    is_nums = False
    try:
        # if not numbers, then textwrap
        float(normalized_headings[0][0])
        is_nums = True
    except:
        is_nums = False

    heading_names = []
    if not user_headings == None and not len(user_headings) == 0:
        if len(filtered_headings) != len(user_heading):
            raise Exception("Error, mismatch in user provided heading lengths")

        heading_names = user_headings
    else:
        for idx, heading in enumerate(normalized_headings[0]):
            if heading in filtered_headings:
                if is_times:
                    heading_names.append((xs[0][idx], heading.strftime("%H:%M")))
                elif is_nums:
                    heading_names.append((xs[0][idx], str(heading)))
                elif len(filtered_headings) > 10:
                    heading_names.append((xs[0][idx], textwrap.fill(heading,15)))
                else:
                    heading_names.append((xs[0][idx], textwrap.fill(heading,12)))


    #print(str(heading_names))
    return is_times, is_nums, xs, normalized_headings, filtered_headings, heading_names

class TemplateEngine:

    def __init__(self, version, inputfile, outputdir, cached_values_regex = None, build_outputs = True):
        if os.path.dirname(inputfile) == outputdir:
            raise Exception("refusing to create output in the input directory")

        self.filename = os.path.basename(inputfile)
        self.outdir = outputdir

        if not os.path.exists(outputdir):
            os.makedirs(outputdir)

        logging.debug("Build outputs: " + str(build_outputs))

        if build_outputs:
            if os.path.exists(os.path.join(outputdir, "media")):
                shutil.rmtree(os.path.join(outputdir, "media"))

            if os.path.isdir(os.path.join(os.path.dirname(inputfile), "media")):
                shutil.copytree(os.path.join(os.path.dirname(inputfile), "media"), os.path.join(outputdir, "media"))

            self.outfile = codecs.open(os.path.join(outputdir, self.filename), "w", "utf-8")

        self.config = {"EnergyPlusVersion":version}

        self._csv_files = {}
        self._file = open(inputfile)
        self._shared_formulas = []
        self._open_workbooks = {}
        self.basepath = os.path.dirname(inputfile)
#        self._reference_regex = re.compile("('?(\\[(\\S+)\\])?([^!']*?)'?!)?(\\$?[A-Z]+\\$?[0-9]+)(:(\\$?[A-Z]+\\$?[0-9]+))?")
#        self._reference_regex = re.compile(r"(('?)(\[(\S+?)\])?(\w[^\[!']*?)\2!)?((\$?[A-Z]+\$?[0-9]+)(:(\$?[A-Z]+\$?[0-9]+))?)")
#        self._reference_regex = re.compile(r"(('?)(?(?<=')[^']+|\w+)\2!)?((\$?[A-Z]+\$?[0-9]+)(:(\$?[A-Z]+\$?[0-9]+))?)")
        self._reference_regex = re.compile(r"((')?(\[(.*?)\])?((?(2)([^'])+|\w+))(?(2)')!)?((\$?[A-Z]+\$?[0-9]+)(:(\$?[A-Z]+\$?[0-9]+))?)")
        self._range_regex = re.compile(r"((\$?[A-Z]+)(\$?[0-9]+))(:((\$?[A-Z]+)(\$?[0-9]+)))?")
        self._function_regex = re.compile("([A-Z]+)\\(")
        self._equals_regex = re.compile("(([^><!])=)")
        self._sheet_regex = re.compile("(.*)!(.*)")
        self._cached_values_regex = cached_values_regex
        self._num_charts = 0
        self._cached_values_files = []
        self._found_files = {}
        self._cell_cache = {}

        self.dump_cache_info()

        global engine
        engine = self

    def month_year(self):
        return datetime.date.today().strftime("%B %Y")

    def day_month_year(self):
        return datetime.date.today().strftime("%m %B %Y")

    def findfile(self, filename):
        origfilename = filename

        if os.path.isabs(filename) or filename.startswith("file://"):
            logging.warning("Adjusting absolute path: " + filename)
            logging.debug("Trying adjusted name: " + filename)
            filename = os.path.basename(filename)

            try:
                return self.findfile_impl(filename)
            except:
                try:
                    o2 = origfilename.replace('\\', '/')
                    filename = os.path.join(os.path.basename(os.path.dirname(o2)), os.path.basename(o2))
                    logging.debug("Trying adjusted name: " + filename)
                    return self.findfile_impl(filename)
                except:
                    raise FileLookupException(origfilename)
        else:
            return self.findfile_impl(filename)


    def findfile_impl(self, filename):
        if filename in self._found_files:
            return self._found_files[filename]


        try1 = os.path.join(self.basepath, filename)
        try2 = os.path.join(os.path.dirname(self.basepath), filename)
        try3 = os.path.join(self.outdir, filename)
        try4 = os.path.join(os.path.dirname(self.outdir), filename)


        if os.path.isfile(try1):
            logging.debug("Found file: '" + filename + "' at '" + str(try1) + "'")
            self._found_files[filename] = try1
            return try1
        elif os.path.isfile(try2):
            logging.debug("Found file: '" + filename + "' at '" + str(try2) + "'")
            self._found_files[filename] = try2
            return try2
        elif os.path.isfile(try3):
            logging.debug("Found file: '" + filename + "' at '" + str(try3) + "'")
            self._found_files[filename] = try3
            return try3
        elif os.path.isfile(try4):
            logging.debug("Found file: '" + filename + "' at '" + str(try4) + "'")
            self._found_files[filename] = try4
            return try4
        else:
            raise FileLookupException(filename)



    def STDEV(self, wb, sheet, *values):
        workingvals = non_empty_floats(values)
        s = sum(workingvals)
        avg = s / len(workingvals)

        top_sum = 0
        for i in workingvals:
            top_sum += (i - avg) ** 2

        return math.sqrt(top_sum / (len(workingvals) - 1))


    def SUM(self, wb, sheet, *values):
        return sum(non_empty_floats(values))

    def AVERAGE(self, wb, sheet, *values):
        working_vals = non_empty_floats(values)
        return sum(working_vals) / len(working_vals)

    def COUNTIF(self, wb, sheet, values, value):

        working_vals = non_empty_floats(values)

        retval = None
        if value[0:2] == ">=":
            #logging.debug("COUNTIF >= (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v >= float(value[2:]) else 0 for v in working_vals)

        if value[0:2] == "<=":
            #logging.debug("COUNTIF <= (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v <= float(value[2:]) else 0 for v in working_vals)

        if value[0:2] == "==":
            #logging.debug("COUNTIF == (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v == float(value[2:]) else 0 for v in working_vals)

        if value[0:1] == ">":
            #logging.debug("COUNTIF > (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v > float(value[1:]) else 0 for v in working_vals)

        if value[0:1] == "<":
            #logging.debug("COUNTIF < (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v < float(value[1:]) else 0 for v in working_vals)

        if value[0:1] == "=":
            #logging.debug("COUNTIF = (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v == float(value[1:]) else 0 for v in working_vals)

        if value[0:2] == "<>":
            #logging.debug("COUNTIF <> (" + str(working_vals) + ") " + str(value))
            retval = sum(1 if v != float(value[1:]) else 0 for v in working_vals)

        if retval != None:
            #logging.debug("COUNTIF: " + str(retval))
            return retval


        return sum(1 if float(value) == v else 0 for v in working_vals)

    def CONCATENATE(self, wb, sheet, *values):
        return "".join(values)

    def format_number(self, value, f):
        try:
            if f == "General":
                if isinstance(value, int):
                    return value
                elif isinstance(value, float):
                    return "{:.6f}".format(value)
                else:
                    return value
            if f == '0.000000_);\(0.000000\)':
                return "{:.6f}".format(value)
            if f == "0.00000":
                return "{:.5f}".format(value)
            elif f == "0.00" or f == "0.00_);\(0.00\)":
                return "{:.2f}".format(value)
            elif f == "0.000":
                return "{:.3f}".format(value)
            elif f == "0.0000":
                return "{:.4f}".format(value)
            elif f == "0.000000":
                return "{:.6f}".format(value)
            elif f == "0.0":
                return "{:.1f}".format(value)
            elif f == "0.00E+00":
                return "{:.2E}".format(value)
            elif f == "0":
                return str(int(round(value)))
            elif f == "0.0%":
                return "{:.1f}%".format(value * 100)
            elif f == "0.00%":
                return "{:.2f}%".format(value * 100)
            elif f == "0.000%":
                return "{:.3f}%".format(value * 100)
            elif f == "0.0000%":
                return "{:.4f}%".format(value * 100)
            elif f == "0.00000%":
                return "{:.5f}%".format(value * 100)
            elif f == "d-mmm-yy":
                return value.strftime("%d-%b-%y")
            elif f == "mm-dd-yy":
                return value.strftime("%m-%d-%y")
            elif f == "d-mmm":
                return value.strftime("%d-%b")
            elif f == "h:mm":
                return value.strftime("%H:%M")
            elif f == "h:mm AM/PM":
                return value.strftime("%H:%M")
            elif f == '#,##0':
                return "{:,}".format(int(value))
            elif f == '#,##0.0':
                return "{:,.1f}".format(value)
            elif f == '#,##0.00000':
                return "{:,.5f}".format(value)
            elif f == '#,##0.00000000':
                return "{:,.8f}".format(value)
            elif f == '_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,}".format(abs(int(v)))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

            elif f == '_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,.2f}".format(abs(v))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

            elif f == '_(* #,##0.000_);_(* \(#,##0.000\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,.3f}".format(abs(v))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

            elif f == '_(* #,##0.0_);_(* \(#,##0.0\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,.1f}".format(abs(v))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

            elif f == '_(* #,##0.0000_);_(* \(#,##0.0000\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,.4f}".format(abs(v))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

            elif f == '_(* #,##0.000000_);_(* \(#,##0.000000\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,.6f}".format(abs(v))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

            elif f == '_(* #,##0.0000000_);_(* \(#,##0.0000000\);_(* "-"??_);_(@_)':
                v = float(value)
                negative = v < 0
                valstr = "{:,.7f}".format(abs(v))
                if negative:
                    valstr = "(" + valstr + ")"

                return valstr

        except Exception as e:
            logging.warning("Unable to handle formatting request (" + f + ") for '" + str(value)+ "' '" + str(e) + "'")
            return value


        raise Exception("Unhandled number format: '" + f + "'")


    def TEXT(self, wb, sheet, value, f):
        return self.format_number(value, f)

    def ISNUMBER(self, wb, sheet, val):
        if isinstance(val, bool):
            return False

        try:
            float(val)
            return True
        except:
            return False

    def ABS(self, wb, sheet, param):
        return abs(param)

    def OR(self, wb, sheet, *params):
        return any(params)

    def AND(self, wb, sheet, *params):
        return all(params)

    def NOT(self, wb, sheet, val):
        return not bool(val)


    def MAX(self, wb, sheet, *values):
        vals = non_empty_floats(values)
        if len(vals) == 0:
            return None
        return max(vals)

    def MIN(self, wb, sheet, *values):
        vals = non_empty_floats(values)
        if len(vals) == 0:
            return None
        return min(vals)

    def INDEX(self, wb, sheet, values, row):
        try:
            return float(values[row-1])
        except:
            return values[row-1]

    def MATCH(self, wb, sheet, value, values, matchtype=0):
        if matchtype == 1:
            index = 0;
            for val in values:
                index += 1

                try:
                    if val >= value:
                        # logging.debug("MATCH: " + str(val) + " to " + str(value) + ": " + str(index))
                        return index-1
                except:
                    try:
                        if float(val) >= float(value):
                            # logging.debug("MATCH: " + str(val) + " to " + str(value) + ": " + str(index))
                            return index-1
                    except:
                        logging.warning("Error comparing values: '" + str(val) + "' '" + str(value) + "'")


            return index

        if matchtype == 0:
            for idx, val in enumerate(values):
                if val == value or str(val) == str(value):
                    return idx + 1

        return False




    def MID(self, wb, sheet, value, start, num):
        return value[start-1:start+num-1]

    def IF(self, wb, sheet, condition, iftrue, iffalse):
        if condition:
            return iftrue
        else:
            return iffalse


    def dump_cache_info(self):
        logging.info("RESOLVE: " + str(self.RESOLVE.cache_info()))
        logging.info("resolve_range: " + str(self.resolve_range.cache_info()))
        logging.info("resolve: " + str(self.resolve.cache_info()))
        logging.info("resolve_chart_range: " + str(self.resolve_chart_range.cache_info()))
        logging.info("resolve_chart_value_range: " + str(self.resolve_chart_value_range.cache_info()))

    def resolve_csv(self, filename, reference):
        min_col, min_row, max_col, max_row = worksheet.range_boundaries(reference)

        #print("resolving: " + filename + " " + reference)

        #print(str(min_col) + " " + str(min_row) + " " + str(max_col) + " " + str(max_row))

        values = []

        csvfile = None

        if filename in self._csv_files:
            csvfile = self._csv_files[filename]
        else:
            with open(self.findfile(filename), 'r') as csvfile:
                reader = csv.reader(csvfile, delimiter=',')

                rows = []
                for row in reader:
                    rows.append(row)

                csvfile = rows
                self._csv_files[filename] = csvfile


        for row in csvfile[min_row-1:max_row]:
            values.extend(row[min_col-1:max_col])

#        logging.debug("resolve_csv: " + filename + " " + reference + " " + str(values))
        return values

    def do_col_offset(self, col, coloffset):
        if coloffset == 0 or col == None or col[0] == "$":
            return col

        return worksheet.get_column_letter(worksheet.column_index_from_string(col) + coloffset)

    def do_row_offset(self, row, rowoffset):
        if rowoffset == 0 or row == None or row[0] == "$":
            return row

        return str(int(row) + rowoffset)

    def SQRT(self, wb, sheet, value):
        return math.sqrt(value)

    def LN(self, wb, sheet, value):

        try:
            return math.log(value)
        except:
            logging.warning("Error calculating LN returning NaN of: " + str(value))
            return float('nan')


    @functools.lru_cache(maxsize=65536)
    def RESOLVE(self, wb, sheet, coloffset, rowoffset, text):
        if rowoffset != 0 or coloffset != 0:
            orig = text
            match = self._reference_regex.match(text)
            rrange = match.group(7)
            rangematch = self._range_regex.match(rrange)

            begincolstr = self.do_col_offset(rangematch.group(2), coloffset)
            beginrowstr = self.do_row_offset(rangematch.group(3), rowoffset)

            endcolstr = self.do_col_offset(rangematch.group(6), coloffset)
            endrowstr = self.do_row_offset(rangematch.group(7), rowoffset)

            newrange = begincolstr + beginrowstr
            if endcolstr:
                newrange += ":" + endcolstr + endrowstr

            text = text.replace(rrange, newrange)

            # print("Updated range from: " + orig + " to: " + text)



        retvals = self.resolve_reference(wb, sheet, text)
        # print("Resolved 2 : " + str(sheet) + ":" + text + " to: " + str(retvals))
        if isinstance(retvals, list):
            if len(retvals) == 1:
                val = retvals[0]

                if isinstance(val, bool):
                    return val

                try:
                    return float(val)
                except:
                    return val
            elif len(retvals) == 0:
                return None

        if retvals == "":
            return 0.0

        return retvals

    def resolve_reference(self, wb, sheet, text):
        match = self._reference_regex.match(text)

        filereference = match.group(4)
        sheetname = match.group(5)
        rrange = match.group(7)

        # print("Resolving reference: " + str(filereference) + ", " + str(sheetname) + ", " + str(rrange))

        next_wb = wb
        next_sheet = sheet

        if filereference:
            filenamestr = urllib.parse.unquote(wb.external_links[int(filereference) - 1].Target)
            reference = rrange.replace("$", "")

            if filenamestr.lower().endswith(".csv"):
                return self.resolve_csv(filenamestr, reference)
            else:
                # print("resolving: " + filenamestr + " " + sheetstr + " " + reference)
                next_wb = self.get_workbook(self.findfile(filenamestr))

        if sheetname:
            next_sheet = next_wb[sheetname]


        retvals = []

        #print("range: " + rrange);
        for row in next_sheet.iter_rows(rrange):
            for cell in row:
                # print("Resolving cell (" + cell.parent.title + "!" + str(cell.column) + str(cell.row) + "): " + str(cell.value) + "; " + str(cell));
                retvals.append(self.resolve(next_wb, next_sheet, cell))

        # print("Resolved range : " + str(sheet) + ":" + rrange + " to: " + str(retvals))

        if isinstance(retvals, list):
            if len(retvals) == 1:
                return retvals[0]
            elif len(retvals) == 0:
                return None

        return retvals


    def get_shared_formula(self, sheet, cell):
        # print("We found a shared formula")
        colindex = worksheet.column_index_from_string(cell.column)
        rowindex = cell.row

        for val in self._shared_formulas:
            foundsheet, min_col, min_row, max_col, max_row, parentcellcol, parentcellrow, formula = val

            if sheet == foundsheet and min_col <= colindex and min_row <= rowindex and max_col >= colindex and max_row >= rowindex:
                return formula, colindex - parentcellcol, rowindex - parentcellrow


        for key in sheet.formula_attributes:
            attr = sheet.formula_attributes[key]
            if 'ref' in attr:
                ref = attr['ref']
                min_col, min_row, max_col, max_row = worksheet.range_boundaries(ref)
                if min_col <= colindex and min_row <= rowindex and max_col >= colindex and max_row >= rowindex:
                    # print("Found the function!: " + key)
                    parentcell = sheet.cell(key)
                    formula = parentcell.value[1:]

                    parentcellcol = worksheet.column_index_from_string(parentcell.column)
                    parentcellrow = parentcell.row

                    coloffset = colindex - parentcellcol
                    rowoffset = rowindex - parentcellrow
                    self._shared_formulas.append((sheet, min_col, min_row, max_col, max_row, parentcellcol, parentcellrow, formula))

                    return formula, coloffset, rowoffset

                    #print("offset: (" + str(coloffset) + ", " + str(rowoffset) + ")")



    @functools.lru_cache(maxsize=65536)
    def resolve(self, wb, sheet, cell):
        if cell in self._cell_cache:
            return self._cell_cache[cell]

        # '[1]Cases 600'!$I$24

        value = cell.value

        if value == None:
            return ""

        strvalue = str(value)

        #print("resolving: " + strvalue)

        if strvalue[0] == '=':
            formula = strvalue[1:]
            coloffset = 0
            rowoffset = 0

            if len(strvalue) == 1:
                formula, coloffset, rowoffset = self.get_shared_formula(sheet, cell)


            yo = self._function_regex.sub("self.\\1(wb, sheet, ", formula)
            yo = self._reference_regex.sub("self.RESOLVE(wb, sheet, " + str(coloffset) + ", " + str(rowoffset) + ", \"\\g<0>\")", yo)
            yo = self._equals_regex.sub("\\2==", yo)
            yo = yo.replace("^", "**")
            yo = yo.replace("FALSE", "False")
            yo = yo.replace("TRUE", "True")

            # print("Orig: '" + strvalue + "'  Fixed up: '" + yo + "'")


            try:
                if yo.find("#REF") != -1:
                    logging.warning("Found invalid reference, returning NaN: '" + yo + "' orig: '" + strvalue)
                    return float('nan')

                try:


                    val = eval(yo)
                    # logging.debug("Evaling: (" + str(cell) + ") '" + yo + "' orig: '" + strvalue + " val: " + str(val))
                    if isinstance(val, list):
                        if len(val) == 1:
                            return val[0]
                        elif len(val) == 0:
                            return None

                    return val
                except FileLookupException as e:
                    if self._cached_values_regex.match(e.filename):
                        if not e.filename in self._cached_values_files:
                            logging.warning("Returning cached values for file: " + e.filename)
                            self._cached_values_files.append(e.filename)

                        return float(cell._cached_value)
                    else:
                        raise

                except ZeroDivisionError as e:
                    return 0

            except:
                logging.error("Error while evaling: '" + yo + "' orig: '" + strvalue)
                raise
        else:
            return value

    def reset_sheet_caches(self):
        self._cell_cache = {}

    def check_load_sheet(self, filename, sheet):
        self.load_sheet(filename, sheet, True, False)
        self.load_sheet(filename, sheet, False, True)

    def load_sheet(self, filename, sheet, force_write=False, check_value=False):
        logging.info("Pre-loading sheet: " + filename + " " + sheet)
        cache_filename = os.path.join(self.outdir, filename + "!" + sheet + ".cache")
        cache_file = None
        is_read = False
        if os.path.isfile(cache_filename) and not force_write:
            cache_file = open(cache_filename,'r')
            is_read = True
        else:
            cache_file = open(cache_filename,'w')


        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[sheet]
        dimensions = sheet_ranges.calculate_dimension()


        if is_read:
            for idx, row in enumerate(sheet_ranges.iter_rows(dimensions)):
                logging.debug("Loading row from cache: " + str(idx))
                row_array = []
                for cell in row:
                    next_line = cache_file.readline()
                    next_line = next_line[:len(next_line)-1]
                    val = None
                    if next_line != "None":
                        try:
                            val = float(next_line)
                        except:
                            if next_line == "False":
                                val = False
                            elif next_line == "True":
                                val = True
                            else:
                                val = next_line.replace("\\n", "\n")

                    if check_value:
                        logging.debug("Checking value: " + str(cell))
                        if self._cell_cache[cell] != val:
                            logging.error("Error loading value: " + str(cell))


                    self._cell_cache[cell] = val
                # logging.debug("Cache size: " + str(len(self._cell_cache)))

                # logging.debug("Loading row: " + str(idx) + " finished")
        else:
            for idx, row in enumerate(sheet_ranges.iter_rows(dimensions)):
                logging.debug("Loading row: " + str(idx))
                row_array = []
                for cell in row:
                    val = self.resolve(wb, sheet_ranges, cell)
                    cache_file.write(str(val).replace("\n", "\\n") + "\n")
                    self._cell_cache[cell] = val

                logging.debug("Loading row: " + str(idx) + " finished")
                self.dump_cache_info()

        logging.info("Pre-loading sheet: " + filename + " " + sheet + " finished")


    @functools.lru_cache(maxsize=65536)
    def resolve_chart_range(self, wb, sheet, r, format_markdown=False, return_grid=False):
        return self.resolve_range(wb, sheet, r, format_markdown, return_grid)

    @functools.lru_cache(maxsize=65536)
    def resolve_chart_value_range(self, wb, sheet, r, format_markdown=True, return_grid=False):
        valuestr = self.resolve_chart_range(wb, sheet, r, format_markdown, return_grid)
        values = []
        breaks = []
        begin = 0
        for idx, s in enumerate(valuestr):
            if s == None or str(s).strip() == "":
                values.append(0.0)
            else:
                if isinstance(s, float):
                    values.append(s)

                elif isinstance(s, bool) or (isinstance(s, str) and (s.lower() == "true" or s.lower() == "false")):
                    values.append(bool(s))

                else:
                    try:
                        values.append(float(s))
                    except:
                        breaks.append((begin, idx))
                        begin = idx + 1
                        values.append(s)

        breaks.append((begin, len(valuestr)))
        return values, breaks


    @functools.lru_cache(maxsize=65536)
    def resolve_range(self, wb, sheet, r, format_markdown=True, return_grid=False):
        retvals = []
        fretvals = []
        # print("Resolving range: " + str(r))
        for currange in r.replace("(", "").replace(")", "").split(","):
            sheet_match = self._sheet_regex.match(currange)

            if sheet_match:
                currange = sheet_match.group(2)
                sheetname = sheet_match.group(1)
                if sheetname[0] == "'":
                    sheetname = sheetname[1:len(sheetname)-1]

                sheet = wb[sheetname]

            # print("Resolving range (" + sheet.title + "!" + currange + ") orig: " + r)
            try:
                for row in sheet.iter_rows(currange.strip()):
                    row_array = []
                    frow_array = []
                    for cell in row:
                        try:
                            if return_grid:
                                #logging.debug("Cell number format: " + str(cell.number_format) + " style: " + str(cell.style) + " font " + str(cell.font) + " alignment: " + str(cell.alignment))
                                value = self.resolve(wb, sheet, cell)

                                if format_markdown:
                                    if cell.number_format != None:
                                        value = self.format_number(value, cell.number_format)

                                    if isinstance(value, str):
                                        value = value.strip()
                                        value = re.sub('(( )+-){2,}\s?>$', '', value)
                                        value = re.sub('{{(.*?)}}', eval_replacement, value)


                                row_array.append(value)
                                frow_array.append(cell)
                            else:
                                retvals.append(self.resolve(wb, sheet, cell))
                        except:
                            logging.error("Error while resolving (" + cell.parent.title + "!" + str(cell.column) + str(cell.row) + ")");
                            raise
                    if return_grid:
                        retvals.append(row_array)
                        fretvals.append(frow_array)
            except:
                logging.error("Error trying to resolve range: (" + sheet.title + ") " + currange)
                raise

        # print("Resolved range (" + sheet.title + "!" + currange + ") to: '" + str(retvals) + "'")

        if return_grid and format_markdown:
            return retvals, fretvals
        else:
            return retvals



    def my_write(self, s):
        if s != None and s.strip() != "None":
            if self.outfile:
                self.outfile.write(s)
            else:
                sys.stdout.write(s)

    def write_chart(self, chart_type, chart_name, chart_title, xaxis, yaxis, filename, tabname, superrange, seriesdata, notes, headings=[], styles=[], do_filter=True, force_times=False):
        self.my_write(self.create_chart(chart_type, chart_name, chart_title, xaxis, yaxis, filename, tabname, superrange, seriesdata, notes, headings, styles=styles, do_filter=do_filter, force_times=force_times))

    def get_bar_color(self, index, num_colors):
        return cm.jet(1.*index/num_colors)

#        colors = ['b', 'g', 'r', 'c', 'm', 'y', 'k',
#                  (.90,.60,  0),
#                  (  0,.60,.50),
#                  (.95,.90,.25),
#                  (  0,.45,.70),
#                  (.80,.40,  0),
#                  (.80,.60,.70),
#                  (.45,.30,  0),
#                  (.17,.35,.45),
#                  (  0,.30,.25),
#                  (.47,.45,.13),
#                  (  0,.22,.35),
#                  (.40,.20,  0),
#                  (.40,.30,.35)]
#
#        return colors[index]

    def get_line_style(self, chart_type, index, styles):
        colors = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
        markers = ['o', 'v', '^', '<', '>', '1', '2', '3', '4', 's', 'p', '*', 'h', 'H', '+', 'x', 'D', 'd', '|', '_']

        if index >= len(styles):
            if chart_type == "Scatter":
                return str(colors[index % len(colors)]) + str(markers[index % len(markers)])
            else:
                return str(colors[index % len(colors)]) + str(markers[index % len(markers)]) + "-"
        else:
            return styles[index]

    def render_slab_data(self, filename):
        return "\n    " + "    ".join(open(self.findfile(filename),'r').readlines()[2:23]) + "\n"

    def percent_abs_range(self, filename, sheet, r):
        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[sheet]
        values = self.resolve_range(wb, sheet_ranges, r)
        absvalues = [abs(x) for x in values]
        return "{:.2f}%".format(min(absvalues) * 100) + " to " + "{:.2f}%".format(max(absvalues) * 100)

    def max_percent_abs(self, filename, sheet, r):
        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[sheet]
        values = self.resolve_range(wb, sheet_ranges, r)
        absvalues = [abs(x) for x in values]
        #logging.debug("Values: " + str(values))
        #logging.debug("ABS Values: " + str(absvalues))

        return "{:.2f}%".format(max(absvalues) * 100)

    def percent_range(self, filename, sheet, r):
        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[sheet]
        values = self.resolve_range(wb, sheet_ranges, r)
        return "{:.2f}%".format(min(values) * 100) + " to " + "{:.2f}%".format(max(values) * 100)

    def create_chart(self, chart_type, chart_name, chart_title, xaxis, yaxis, filename, tabname, superrange, seriesdata, notes, headings=[], styles=[], do_filter=True, force_times=False):
        logging.debug("Creating chart: " + chart_name)
        chart_title = re.sub('{{(.*?)}}', eval_replacement, chart_title)
        self.dump_cache_info()
        output = ""

        outdir = os.path.join(self.outdir, "generated_media")
        outfilename = os.path.join(outdir, str(self._num_charts).zfill(3) + "_" + chart_name) + ".svg"
        outfile = os.path.relpath(outfilename, self.outdir)



        self._num_charts += 1

#        if os.path.isfile(outfilename):
#            logging.warning("Not re-creating existing chart: " + outfilename)
#            return output


        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[tabname]

        data_series_names = []
        data_series_headings = []
        data_series_values = []
        dataranges_split = -1

        for index, dataranges in enumerate(seriesdata.split("#")):
            if index > 0:
                dataranges_split = len(data_series_names)

            for series in dataranges.split(";"):
                if series != "":
                    # print("matching: " + series)
                    match = re.match("\\((.*?),\\(?([^()]*)\\)?,\\(?([^()]*)\\)?,[0-9]+\\)", series)

                    if match == None:
                        logging.error("Error matching data series string: " + str(series))

                    data_series_names.append(match.group(1))
                    data_series_headings.append(match.group(2))
                    data_series_values.append(match.group(3))

        is_times, is_nums, xs, headings, filtered_headings, heading_names = None, None, None, None, None, None

        try:
            is_times, is_nums, xs, headings, filtered_headings, heading_names = normalize_headings([self.resolve_chart_range(wb, sheet_ranges, ds) for ds in data_series_headings], headings)
        except:
            logging.error("Error while normalizing headings: headings: '" + str(data_series_headings) + "'")
            raise


        plotsize = len(heading_names)

        fig, ax = plt.subplots()


        index = np.arange(len(heading_names))
        bar_width = 1 / (len(data_series_values) + 2)
        opacity = 1

        # print("rnage: " + str(range(0, len(data_series_values))))

        ax2 = None
        if dataranges_split > -1:
            ax2 = ax.twinx()

        notelocations = []
        lastbreaks = []
        labels = []
        for i in range(0, len(data_series_values)):

            values, breaks = self.resolve_chart_value_range(wb, sheet_ranges, data_series_values[i])

            if len(breaks) < len(lastbreaks):
                breaks = lastbreaks
            else:
                lastbreaks = breaks

            plotsize = breaks[0][1] - breaks[0][0]

            label = ""

            if (data_series_names[i][0] == "'" or data_series_names[i][0] == '"') and (data_series_names[i][-1:] == "'" or data_series_names[i][-1] == '"'):
                # look for literal quoted label name, not just a reference
                label = data_series_names[i][1:len(data_series_names[i])-1]
            else:
                label = self.resolve_range(wb, sheet_ranges, data_series_names[i])[0]

            labels.append(label)

            hatch = ''


            if chart_type == "ColumnClustered":
                # logging.debug("Col " + str(i) + " vals " + str(values))
                plt.bar(index + i * bar_width, values, bar_width,
                        alpha = opacity,
                        label = label,
                        color= self.get_bar_color(i, len(data_series_values)),
                        hatch=hatch
                        )
            elif chart_type == "ScatterLines" or chart_type == "Combo" or chart_type == "ScatterLinesNoMarkers" or chart_type == "Scatter" or chart_type == "ScatterSmooth":
                plotloc = None
                if ax2 and i >= dataranges_split:
                    # print("going to ax2!")
                    plotloc = ax2
                else:
                    plotloc = ax

                for idx, b in enumerate(breaks):
                    x = xs[i][b[0]:b[1]]
                    vals = values[b[0]:b[1]]
                    heads = headings[i][b[0]:b[1]]

                    if i == 0:
                        notelocations.append((x[int(len(x)/2)], vals[int(len(x)/2)]))


                    filteredx = []
                    filteredvals = []


                    if do_filter:
                        for idx2, head in enumerate(heads):
                            if head in filtered_headings:
                                filteredx.append(x[idx2])
                                filteredvals.append(vals[idx2])
                    else:
                        filteredx = x
                        filteredvals = vals


                    if any(v != 0.0 for v in vals):
#                        if len(x) < len(vals):
#                            logging.warning("Truncating values for chart: [" + filename + "]'" + chart_name + "'")
#                        elif len(x) > len(vals):
#                            logging.error("Error, fewer values than headers!")
#                            raise Exception("Not enough values provided vals: '" + str(vals) + "' headers: '" + str(x) + "'")

                        if len(filteredx) == 0:
                            print("xs: " + str(x) + " vals: " + str(vals) + " heads: " + str(heads))
                            print("filtered xs: " + str(filteredx) + " vals: " + str(filteredvals) + " heads: " + str(heads))
                            print("Filtered headings: '" + str(filtered_headings) + "'")

#                        print("xs: " + str(x) + " vals: " + str(vals) + " heads: " + str(heads))
#                        print("filtered xs: " + str(filteredx) + " vals: " + str(filteredvals) + " heads: " + str(heads))
#                        print("Filtered headings: '" + str(filtered_headings) + "'")

                        name = label
                        if idx > 0:
                            name = "_" + str(name)

                        # logging.debug("Plotting '" + str(name) + "': '" + str(filteredx) + "' '" + str(filteredvals) + "'")
                        line_style = self.get_line_style(chart_type, i, styles)

                        if isinstance(line_style, tuple) or isinstance(line_style, list):
                            # print("Line Style Chosen: " + line_style + " for chart type: " + chart_type)
                            line = plotloc.plot(filteredx,
                                     filteredvals,
                                     color=line_style[0],
                                     linestyle=line_style[1],
                                     marker=line_style[2],
                                     label=name
                                    )
                        else:
#                            print("Line Style Chosen: " + str(line_style) + " for chart type: " + chart_type)
                            line = plotloc.plot(filteredx,
                                     filteredvals,
                                     line_style,
                                     label=name
                                    )

            else:
                raise Exception("Unknown chart_type: " + chart_type)


        if xaxis != None and len(xaxis) > 0:
            plt.xlabel(xaxis, size='small')

        if ax2:
            ax.set_ylabel(yaxis[0], size='small')
            ax2.set_ylabel(yaxis[1], size='small')
        else:
            if yaxis != None and len(yaxis) > 0:
                plt.ylabel(yaxis, size='small')


        plt.title(chart_title, size='small')
        plt.tick_params(axis='both', which='major', labelsize='xx-small')

#        if plotsize == len(headings):
        if force_times:
            xvals = []
            xlabels = []
            for x in heading_names[:plotsize]:
                dt = datetime.datetime(2000,1,1) + datetime.timedelta(0, round(60*60*24*(float(x[0])-1)))
                # logging.debug("Date: " + str(dt))
                if dt.time().minute == 0:
                    xvals.append(x[0])
                    xlabels.append(dt.strftime("%m-%d %H:%M"))

            # logging.debug("xticks: " + str(xvals))
            # logging.debug("xticklabels: " + str(xlabels))

            plt.xlim(xvals[0], xvals[-1])
            ax.set_xticks(xvals)
            ax.set_xticklabels(xlabels, rotation=45, horizontalalignment='right', size='xx-small')

            # logging.debug("set labels")

        elif not is_times:
            cur_headings = [re.sub('{{(.*?)}}', eval_replacement, heading[1]) for heading in heading_names[:plotsize]]

            if not is_nums and len(cur_headings) > 9:
                logging.debug("xticks labels: " + str(cur_headings) + " " + str(len(cur_headings)))
                ax.set_xticklabels(cur_headings, rotation=45, horizontalalignment='right', size='xx-small')
                plt.xlim(0, len(cur_headings))
                plt.xticks(index + .5, cur_headings)
            else:
                if not is_nums:
                    logging.debug("xticks labels: " + str(cur_headings) + " " + str(len(cur_headings)))
                    index = np.arange(plotsize)
                    ax.set_xticklabels(cur_headings, horizontalalignment='center', size='xx-small')
                    plt.xlim(0, len(cur_headings))
                    plt.xticks(index + .5, cur_headings)
        else:
            xvals = []
            xlabels = []
            for idx, x in enumerate(heading_names[:plotsize]):
                if x[0] % 60 == 0:
                    xvals.append(x[0])
                    xlabels.append(x[1])

            ax.set_xticks(xvals)
            ax.set_xticklabels(xlabels, size='xx-small')


        plt.subplots_adjust(bottom=0.2)

        box = ax.get_position()
        ax.set_position([box.x0, box.y0 + box.height * 0.1, box.width, box.height * 0.8])

        if ax2:
            ax2.set_position([box.x0, box.y0 + box.height * 0.1, box.width, box.height * 0.8])



        lines, labels = ax.get_legend_handles_labels()
        labels = [re.sub('{{(.*?)}}', eval_replacement, label) for label in labels]

        legendcolumns = 4
        if any(len(l) > 30 for l in labels):
            legendcolumns = 2
        elif any(len(l) > 20 for l in labels):
            legendcolumns = 3

        if ax2:
            lines2, labels2 = ax2.get_legend_handles_labels()
            lines += lines2
            labels += labels2

        ax.legend(lines, labels, loc='upper center', bbox_to_anchor=(0.5, -0.25),
                          fancybox=True, shadow=False, ncol=legendcolumns, fontsize='xx-small').get_frame().set_alpha(1)

        notetext = ""
        if len(notes) > 1 and len(notelocations) == len(notes):
            for idx, note in enumerate(notes):
                ax.annotate(note, xy=notelocations[idx],  xycoords='data',
                            xytext=(15, 15), textcoords='offset points',
                            bbox=dict(boxstyle='round,pad=0.2', fc='black', alpha=0.3),
                            arrowprops=dict(facecolor='black', arrowstyle='-'),
                            color='white', size='xx-small',
                            horizontalalignment='right', verticalalignment='center',
                            )
        elif len(notes) > 0:
            notetext = "<strong>Notes:</strong>" + "<br/>".join([re.sub('{{(.*?)}}', eval_replacement, line) for line in notes])

        output += "![" + notetext + "](" + urllib.parse.quote(outfile) +")\n\n"

        fig = plt.gcf()
        fig.set_size_inches(8.5,5.5)

        if not os.path.exists(outdir):
            os.makedirs(outdir)

        fig.savefig(outfilename, orientation='landscape', dpi=120)
#        plt.show()
        plt.close()

        logging.debug("Creating chart: " + chart_name + " finished: " + outfilename)

        return output







    def create_paragraphs(self, filename, column):
        result = ""
        f = yaml.load(codecs.open(os.path.join(self.basepath, filename), "r", "utf-8"))

        for record in f:
            if column in record:
                result += record[column] + "\n"

        return result

    def create_table_from_yaml(self, filename, columns):
        f = yaml.load(codecs.open(os.path.join(self.basepath, filename), "r", "utf-8"))

        rows = []
        for record in f:
            row = []
            for col in columns:
                if col in record:
                    row.append(record[col])
                else:
                    row.append(None)

            rows.append(row)

        return self.create_table(rows, columns)

    def get_workbook(self, filename):
        if filename in self._open_workbooks:
            return self._open_workbooks[filename]
        else:
            logging.debug("Loading workbook: " + filename)
            wb = load_workbook(filename = filename)

            # print(wb._named_ranges)

            #for l in wb._external_links:
            #    print(str(l.Id) + ": " + str(l.Target))

            self._open_workbooks[filename] = wb
            logging.debug("Loading workbook: " + filename + " finished")
            return wb

    def get_cell(self, filename, reference):
        return self.RESOLVE(self.get_workbook(self.findfile(filename)), None, 0, 0, reference)

    def format_diff_row(self, filename, reference1, reference2):
        return "{:.4f}".format(self.get_cell(filename, reference1)) + " | " + "{:.4f}".format(self.get_cell(filename, reference2)) + " | " + "{:.2f}%".format(((engine.get_cell(filename, reference2) - engine.get_cell(filename, reference1)) / engine.get_cell(filename, reference1)) * 100)



    def create_table_from_excel(self, filename, sheet, info_rows=[], num_cols=None):
        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[sheet]
        return self.create_table_from_excel_range(filename, sheet, sheet_ranges.calculate_dimension(), info_rows, num_cols)

    def create_table_from_excel_range(self, filename, sheet, datarange, info_rows=[], num_cols=None):
        wb = self.get_workbook(self.findfile(filename))
        sheet_ranges = wb[sheet]
        data, formatting = self.resolve_range(wb, sheet_ranges, datarange, True, True)

        result = self.create_table(data, [], info_rows, formatting, num_cols);
        return result

    def create_table(self, rows, columns, info_rows=[], formatting=[], num_cols=None):
        retval = "\n\n"
        rowstack = []
        formattingstack = []

        for idx, row in enumerate(rows):
            if idx in info_rows:
                if len(rowstack) > 0:
                    retval += "\n\n" + self.create_table_impl(rowstack, columns, formattingstack, num_cols) + "\n"
                    rowstack = []
                    formattingstack = []

                retval += (" ".join(row)).strip() + "  \n"
            else:
                if len(formatting) > 0:
                    formattingstack.append(formatting[idx])

                rowstack.append(row)




        if len(rowstack) > 0:
            retval += "\n\n" + self.create_table_impl(rowstack, columns, formattingstack, num_cols)

        return retval


    def create_table_impl(self, rows, columns, formatting, num_cols):
        if num_cols == None or len(columns) > 0:
            return self.create_table_impl2(rows, columns, formatting)
        else:
            header_col = 0

            for idx, col in enumerate(rows[int(len(rows) / 2)]):
                if col != None and col != '':
                    header_col = idx
                    break

            total_cols = len(rows[0]) - header_col - 1
            retval = ""

            logging.info("Splitting table: header: " + str(header_col) + " total_cols: " + str(total_cols) + " num_cols: " + str(num_cols))

            for split in range(0, int(math.ceil(total_cols / num_cols))):
                split_rows = []
                split_formatting = None

                if formatting != None:
                    split_formatting = []

                split_start = split * num_cols + header_col + 1
                split_end = split_start + num_cols

                logging.info("Splitting table: split: " + str(split_start) + " to " + str(split_end))

                for idx, row in enumerate(rows):
                    new_row = [row[header_col]]
                    new_row.extend(row[split_start:split_end])
                    split_rows.append(new_row)

                    if formatting != None:
                        new_frow = [formatting[idx][header_col]]
                        new_frow.extend(formatting[idx][split_start:split_end])
                        split_formatting.append(new_frow)

                retval += self.create_table_impl2(split_rows, [], split_formatting) + "\n\n"

            return retval



    def create_table_impl2(self, rows, columns, formatting):
        result = ""
        lens = []

        numcols = len(rows[0])

        for _ in range(numcols):
            lens.append(0)

        for idx, val in enumerate(columns):
            lens[idx] = len(val)

        newrows = []

        # Fix up the rows
        for record in rows:
            newrow = []
            for val in record:
                if val:
                    newrow.append(str(val).replace('\n', ' '))
                else:
                    newrow.append(val)
            newrows.append(newrow)

        rows = newrows

        # get their lengths
        for record in rows:
            for idx, val in enumerate(record):
                if val:
                    lens[idx] = max(lens[idx], len(val))

        result += "<table>"

        if len(columns) > 0:
            result += "<thead><tr>"
            for col in columns:
                result += "<th>" + str(col) + "</th>"

            result += "</tr></thead>"

        # Rows
        for rowidx, row in enumerate(rows):
            rowstr = "<tr>"
            rowempty = True
            for colidx, col in enumerate(row):
                strval = ""

                if col != None:
                    strval = str(col).replace("](", "] (").strip()

                if len(strval) > 0:
                    rowempty = False


                tdprops = ""
                skip_cell = False
                if len(formatting) > 0:
                    celldata = formatting[rowidx][colidx]

                    for merge in celldata.parent.merged_cell_ranges:
                        b, e = merge.split(":")
                        bcell = celldata.parent[b]

                        bcol = worksheet.column_index_from_string(bcell.column)
                        brow = bcell.row

                        ecell = celldata.parent[e]
                        ecol = worksheet.column_index_from_string(ecell.column)
                        erow = ecell.row

                        if celldata.column == bcell.column and celldata.row == bcell.row:
                            tdprops += " colspan=" + str(ecol - bcol + 1)
                            tdprops += " rowspan=" + str(erow - brow + 1)
                            logging.debug("Found first cell in merged range: " + str(celldata) + " (" + merge + ")")
                        else:
                            mcol = worksheet.column_index_from_string(celldata.column)
                            mrow = celldata.row


                            if mcol >= bcol and mcol <= ecol and mrow >= brow and mrow <= erow:
                                logging.debug("Found contained (skipped) cell in merged range: " + str(celldata) + " (" + merge + ")")
                                skip_cell = True




                    if len(strval) > 0:
                        if celldata.font.b:
                            strval = "**" + strval + "**"

                    horiz = celldata.style.alignment.horizontal
                    vert = celldata.style.alignment.vertical

                    if horiz == 'right' or horiz == 'left' or horiz == 'center':
                        tdprops += " align='" + horiz + "'"
                    elif horiz == 'general':
                        if re.match("[0-9\\.,%-]", strval):
                            tdprops += " align='right'"


                    if vert == 'top' or vert == 'bottom' or vert == 'center':
                        tdprops += " valign='" + vert + "'"

                    if len(strval) > 0 and strval[0] == '-':
                        tdprops += ' style="white-space:nowrap" '

                if not skip_cell:
                    rowstr += "<td" + tdprops + ">" + strval + "</td>"

            rowstr += "</tr>"

            if not rowempty:
                result += "\n" + rowstr

        result += "</table>"

        return result


    def execute(self):
        in_exec_block = False

        exec_block = ""
        lineno = 1


        try:
            for line in self._file:
                if line.strip() == "```{exec_python}":
                    in_exec_block = True
                elif line.strip() == "```":
                    in_exec_block = False
                    exec(exec_block)
                    exec_block = ""
                elif in_exec_block:
                    exec_block += line
                else:
                    line = re.sub('{{(.*?)}}', eval_replacement, line)

                    self.my_write(line)


                lineno += 1
        except:
            logging.error("Error while processing line (" + str(lineno) + ")")
            raise








